import pygame
import random
import sys
import os
import cv2
from hand_control import HandController
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL = True
except Exception:
    OPENPYXL = False

from datetime import datetime

pygame.init()

# --- KÍCH THƯỚC MÀN HÌNH ---
SCREEN_WIDTH = 900
SCREEN_HEIGHT = 800
screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
pygame.display.set_caption("Game Đập Chuột")

clock = pygame.time.Clock()
FPS = 120

# --- DIFFICULTY / ADAPTIVE SETTINGS (tùy chỉnh để giảm độ khó) ---
DIFFICULTY_PRESETS = {
    "easy":   {"spawn_den": 240, "min_up": 2000, "max_up": 3500, "max_simultaneous": 2},
    "normal": {"spawn_den": 120, "min_up": 1000, "max_up": 2500, "max_simultaneous": 3},
    "hard":   {"spawn_den": 60,  "min_up": 700,  "max_up": 1800, "max_simultaneous": 4},
}

# Mặc định cho người mới phục hồi: easy
DIFFICULTY = "easy"
_spawn_cfg = DIFFICULTY_PRESETS[DIFFICULTY]
SPAWN_DENOM = _spawn_cfg["spawn_den"]         # spawn chance: 1 / SPAWN_DENOM (per frame)
MOLE_UP_MIN_MS = _spawn_cfg["min_up"]
MOLE_UP_MAX_MS = _spawn_cfg["max_up"]
MAX_SIMULTANEOUS_MOLES = _spawn_cfg["max_simultaneous"]

# --- MÀU ---
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
GREEN = (0, 150, 0)
RED = (255, 0, 0)

# --- PHÔNG CHỮ ---
font = pygame.font.SysFont("arial", 64)
small_font = pygame.font.SysFont("arial", 32)

# --- HÀM LOAD ẢNH ---
def load_image(file_name, size=None):
    path = os.path.join('assets', file_name)
    image = pygame.image.load(path).convert_alpha()
    if size:
        image = pygame.transform.scale(image, size)
    return image

BACKGROUND_IMAGE = load_image('background.png', (SCREEN_WIDTH, SCREEN_HEIGHT))
HOLE_IMAGE = load_image('hole.png', (150, 100))
MOLE_IMAGE_UP = load_image('mole.png', (100, 100))
MOLE_IMAGE_DOWN = load_image('hit_mole.png', (100, 100))
HAMMER_IMAGE = load_image('hammer.png', (80, 80))

ANGLES_XLSX = os.path.join(os.path.dirname(__file__), "game_angles.xlsx")
ANGLES_SUMMARY_XLSX = os.path.join(os.path.dirname(__file__), "game_angles_summary.xlsx")

def ensure_angles_xlsx():
    if not OPENPYXL:
        return
    if not os.path.exists(ANGLES_XLSX):
        wb = Workbook()
        ws = wb.active
        # THÊM cột clench_speed
        ws.append(["timestamp","thumb","index","middle","ring","pinky","clench_speed"])
        wb.save(ANGLES_XLSX)

def save_angles_xlsx(angles, clench_speed=0.0):
    if not OPENPYXL:
        return
    ensure_angles_xlsx()
    wb = load_workbook(ANGLES_XLSX)
    ws = wb.active
    ts = datetime.now().isoformat()
    ws.append([
        ts,
        angles.get("thumb",0),
        angles.get("index",0),
        angles.get("middle",0),
        angles.get("ring",0),
        angles.get("pinky",0),
        clench_speed
    ])
    wb.save(ANGLES_XLSX)

# --- summary (per round) functions ---
def ensure_angles_summary_xlsx():
    if not OPENPYXL:
        return
    if not os.path.exists(ANGLES_SUMMARY_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        header = [
            "timestamp", "player", "round",
            "frames_recorded"
        ]
        # add per-finger avg/max/min
        fingers = ["thumb","index","middle","ring","pinky"]
        for f in fingers:
            header += [f + "_avg", f + "_max", f + "_min"]
        # THÊM clench summary
        header += ["clench_avg", "clench_max", "clench_min"]
        ws.append(header)
        wb.save(ANGLES_SUMMARY_XLSX)

def save_angles_summary_xlsx(player, round_no, stats, clench_stats):
    """
    stats: dict finger -> {'count':int,'sum':float,'max':float,'min':float}
    clench_stats: {'count':int,'sum':float,'max':float,'min':float}
    """
    if not OPENPYXL:
        return
    ensure_angles_summary_xlsx()
    wb = load_workbook(ANGLES_SUMMARY_XLSX)
    ws = wb.active
    ts = datetime.now().isoformat()
    fingers = ["thumb","index","middle","ring","pinky"]
    frames = 0
    row = [ts, player, round_no]
    # frames recorded = count for any finger (they should be same)
    for f in fingers:
        frames = max(frames, stats.get(f, {}).get("count", 0))
    row[3:3] = [frames]  # ensure frames at correct index

    # append avg/max/min per finger
    for f in fingers:
        s = stats.get(f, {"count":0,"sum":0.0,"max":0.0,"min":0.0})
        cnt = s.get("count", 0)
        avg = (s.get("sum",0.0)/cnt) if cnt>0 else 0.0
        mx = s.get("max", 0.0) if cnt>0 else 0.0
        mn = s.get("min", 0.0) if cnt>0 else 0.0
        row += [round(avg,1), round(mx,1), round(mn,1)]

    # THÊM clench stats
    c = clench_stats or {"count":0,"sum":0.0,"max":0.0,"min":0.0}
    cnt = c.get("count",0)
    cavg = (c.get("sum",0.0)/cnt) if cnt>0 else 0.0
    cmx = c.get("max",0.0) if cnt>0 else 0.0
    cmn = c.get("min",0.0) if cnt>0 else 0.0
    row += [round(cavg,1), round(cmx,1), round(cmn,1)]

    ws.append(row)
    wb.save(ANGLES_SUMMARY_XLSX)

# --- CLASS MOLE ---
class Mole(pygame.sprite.Sprite):
    def __init__(self, x, y):
        super().__init__()
        self.image_up = MOLE_IMAGE_UP
        self.image_down_hit = MOLE_IMAGE_DOWN
        self.hole_image = HOLE_IMAGE

        self.image = self.hole_image
        self.rect = self.image.get_rect(topleft=(x, y))
        self.is_up = False
        self.hit = False
        self.time_up = 0
        # default up_duration will be set each time show() is called
        self.up_duration = random.randint(MOLE_UP_MIN_MS, MOLE_UP_MAX_MS)
        self.hit_display_time = 300
        self.time_hit = 0

    def show(self):
        if not self.is_up:
            self.is_up = True
            self.hit = False
            self.time_up = pygame.time.get_ticks()
            # set up duration according to current difficulty config
            self.up_duration = random.randint(MOLE_UP_MIN_MS, MOLE_UP_MAX_MS)
            self.image = self.image_up

    def update(self):
        now = pygame.time.get_ticks()
        if self.is_up:
            if self.hit:
                if now - self.time_hit > self.hit_display_time:
                    self.is_up = False
                    self.image = self.hole_image
            elif now - self.time_up > self.up_duration:
                self.is_up = False
                self.image = self.hole_image

    def was_hit(self):
        global score, hit_count
        if self.is_up and not self.hit:
            self.hit = True
            self.image = self.image_down_hit
            self.time_hit = pygame.time.get_ticks()
            score += 10
            hit_count += 1
            return True
        return False

# --- HÀM LƯU LỊCH SỬ ---
def save_score_to_excel(player_name, score, hit_count, accuracy, filename="game_history.xlsx"):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "LichSu"
        ws.append(["Thời gian", "Tên người chơi", "Điểm", "Số lần trúng", "Tỉ lệ phản ứng (%)"])
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([now, player_name, score, hit_count, round(accuracy, 1)])
    wb.save(filename)
    print(f"✅ Đã lưu kết quả của {player_name} vào {filename}")

# --- HÀM VẼ NÚT ---
def draw_button(surface, rect, text, font, bg_color, text_color):
    pygame.draw.rect(surface, bg_color, rect, border_radius=10)
    label = font.render(text, True, text_color)
    surface.blit(label, (rect.x + (rect.width - label.get_width()) // 2,
                         rect.y + (rect.height - label.get_height()) // 2))

# --- HÀM NHẬP TÊN + SỐ LẦN CHƠI ---
def get_player_info():
    # bật hiển thị chuột và text input (hỗ trợ IME)
    pygame.mouse.set_visible(True)
    pygame.key.start_text_input()

    player_name = ""
    num_games = ""
    stage = "name"
    input_active = True

    while input_active:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.key.stop_text_input()
                pygame.quit(); sys.exit()
            elif event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN:
                    if stage == "name" and player_name.strip():
                        stage = "num"
                    elif stage == "num" and num_games.strip().isdigit():
                        input_active = False
                elif event.key == pygame.K_BACKSPACE:
                    if stage == "name":
                        player_name = player_name[:-1]
                    else:
                        num_games = num_games[:-1]
                elif event.key == pygame.K_ESCAPE:
                    pygame.key.stop_text_input()
                    pygame.quit(); sys.exit()
                else:
                    # dùng event.unicode để lấy ký tự nhập
                    if stage == "name" and len(player_name) < 15:
                        player_name += event.unicode
                    elif stage == "num" and event.unicode.isdigit() and len(num_games) < 2:
                        num_games += event.unicode
            elif event.type == pygame.MOUSEBUTTONDOWN:
                # xử lý click nút TIẾP TỤC
                start_button = pygame.Rect(SCREEN_WIDTH // 2 - 100, 400, 200, 60)
                if start_button.collidepoint(event.pos):
                    if stage == "name" and player_name.strip():
                        stage = "num"
                    elif stage == "num" and num_games.strip().isdigit():
                        input_active = False

        screen.blit(BACKGROUND_IMAGE, (0, 0))
        if stage == "name":
            title = "NHAP TEN NGUOI CHOI:"
            current = player_name
        else:
            title = "NHAP SO LAN CHOI:"
            current = num_games

        text = font.render(title, True, WHITE)
        screen.blit(text, (SCREEN_WIDTH // 2 - text.get_width() // 2, 180))

        name_text = small_font.render(current + "|", True, GREEN)
        screen.blit(name_text, (SCREEN_WIDTH // 2 - name_text.get_width() // 2, 300))

        start_button = pygame.Rect(SCREEN_WIDTH // 2 - 100, 400, 200, 60)
        draw_button(screen, start_button, "TIEP TUC", small_font, GREEN, WHITE)

        pygame.display.flip()
        clock.tick(30)

    pygame.key.stop_text_input()
    pygame.mouse.set_visible(False)
    return player_name, int(num_games)

# --- KHỞI TẠO ---
base_x, base_y = 220, 250
x_spacing, y_spacing = 170, 120
mole_positions = [(base_x + i * x_spacing, base_y + j * y_spacing) for j in range(3) for i in range(3)]
moles = [Mole(x, y) for x, y in mole_positions]

hand_controller = HandController()
hand_controller.start_detection()

# --- VÒNG LẶP TOÀN GAME ---
while True:
    player_name, total_rounds = get_player_info()
    round_count = 0

    while round_count < total_rounds:
        score = 0
        hit_count = 0
        total_moles_shown = 0
        game_time = 30
        game_over = False
        start_time = pygame.time.get_ticks()

        # --- INIT ANGLE STATS FOR THIS ROUND ---
        fingers = ["thumb","index","middle","ring","pinky"]
        angle_stats = {f: {"count":0, "sum":0.0, "max":-9999.0, "min":9999.0} for f in fingers}
        # --- MỚI: clench stats ---
        clench_stats = {"count":0, "sum":0.0, "max":-9999.0, "min":9999.0}

        running = True
        while running:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    hand_controller.stop_detection()
                    pygame.quit(); sys.exit()
                if event.type == pygame.MOUSEBUTTONDOWN and game_over:
                    if play_again_rect.collidepoint(pygame.mouse.get_pos()):
                        running = False  # chuyển sang lượt chơi tiếp theo

                # --- Thay đổi độ khó bằng phím 1/2/3 (tùy ý) ---
                if event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_1:
                        DIFFICULTY = "easy"
                    elif event.key == pygame.K_2:
                        DIFFICULTY = "normal"
                    elif event.key == pygame.K_3:
                        DIFFICULTY = "hard"
                    # cập nhật cấu hình sau khi đổi
                    _spawn_cfg = DIFFICULTY_PRESETS[DIFFICULTY]
                    SPAWN_DENOM = _spawn_cfg["spawn_den"]
                    MOLE_UP_MIN_MS = _spawn_cfg["min_up"]
                    MOLE_UP_MAX_MS = _spawn_cfg["max_up"]
                    MAX_SIMULTANEOUS_MOLES = _spawn_cfg["max_simultaneous"]

            hand_position, gesture, cam_frame = hand_controller.get_hand_position()

            # --- update angle stats each frame ---
            angles = hand_controller.last_angles or {}
            # clench speed từ hand_controller
            clench_speed = getattr(hand_controller, "last_clench_speed", 0.0)

            if angles:
                for f in fingers:
                    val = float(angles.get(f, 0.0))
                    s = angle_stats[f]
                    s["count"] += 1
                    s["sum"] += val
                    if val > s["max"]:
                        s["max"] = val
                    if val < s["min"]:
                        s["min"] = val

            # cập nhật clench_stats mỗi frame (nếu có reading)
            if clench_speed is not None:
                c = clench_stats
                c["count"] += 1
                c["sum"] += float(clench_speed)
                if clench_speed > c["max"]:
                    c["max"] = clench_speed
                if clench_speed < c["min"]:
                    c["min"] = clench_speed

            if not game_over:
                elapsed = (pygame.time.get_ticks() - start_time) // 1000
                time_left = game_time - elapsed
                if time_left <= 0:
                    game_over = True
                    acc = (hit_count / total_moles_shown * 100) if total_moles_shown > 0 else 0
                    save_score_to_excel(player_name, score, hit_count, acc)

                for mole in moles:
                    mole.update()

                # giới hạn số moles cùng lúc và dùng SPAWN_DENOM để điều khiển tần suất
                up_count = sum(1 for m in moles if m.is_up)
                if up_count < MAX_SIMULTANEOUS_MOLES and random.randint(1, SPAWN_DENOM) == 1:
                    available = [m for m in moles if not m.is_up]
                    if available:
                        random.choice(available).show()
                        total_moles_shown += 1

                if hand_position and gesture:
                    for mole in moles:
                        if mole.rect.collidepoint(hand_position):
                            mole.was_hit()

            screen.blit(BACKGROUND_IMAGE, (0, 0))
            for mole in moles:
                screen.blit(mole.hole_image, mole.rect)
            for mole in moles:
                if mole.is_up:
                    screen.blit(mole.image, mole.rect)

            score_text = small_font.render(f"Score: {score}", True, BLACK)
            time_text = small_font.render(f"Time: {max(0, game_time - (pygame.time.get_ticks() - start_time)//1000)}s", True, BLACK)
            screen.blit(score_text, (10, 10))
            screen.blit(time_text, (SCREEN_WIDTH - time_text.get_width() - 10, 10))

            if hand_position:
                screen.blit(HAMMER_IMAGE, (hand_position[0]-30, hand_position[1]-30))

            if game_over:
                # save per-round summary
                save_angles_summary_xlsx(player_name, round_count + 1, angle_stats, clench_stats)
                # also optionally save last frame angles + clench speed as row
                if angles:
                    save_angles_xlsx(angles, clench_speed)

                pygame.mouse.set_visible(True)
                overlay = pygame.Surface((SCREEN_WIDTH, SCREEN_HEIGHT), pygame.SRCALPHA)
                overlay.fill((0, 0, 0, 180))
                screen.blit(overlay, (0, 0))

                acc = (hit_count / total_moles_shown * 100) if total_moles_shown > 0 else 0
                texts = [
                    font.render("GAME OVER", True, RED),
                    small_font.render(f"Ten: {player_name}", True, WHITE),
                    small_font.render(f"Diem: {score}", True, WHITE),
                    small_font.render(f"So lan nam tay: {hit_count}", True, WHITE),
                    small_font.render(f"Ti le phan ung: {acc:.1f}%", True, WHITE)
                ]
                for i, t in enumerate(texts):
                    screen.blit(t, (SCREEN_WIDTH//2 - t.get_width()//2, 200 + i*60))

                play_again_rect = pygame.Rect(SCREEN_WIDTH//2 - 100, 600, 200, 60)
                draw_button(screen, play_again_rect, "LUOT TIEP", small_font, GREEN, WHITE)

            if cam_frame is not None:
                cam_h, cam_w = 150, 180
                cam_frame = cv2.resize(cam_frame, (cam_w, cam_h))
                cam_frame = cv2.cvtColor(cam_frame, cv2.COLOR_BGR2RGB)
                cam_surface = pygame.surfarray.make_surface(cam_frame.swapaxes(0, 1))
                screen.blit(cam_surface, (SCREEN_WIDTH - cam_w - 10, SCREEN_HEIGHT - cam_h - 10))

            pygame.display.flip()
            clock.tick(FPS)

        round_count += 1

    # --- Hết số lần chơi ---
    screen.fill(BLACK)
    msg = font.render("DA HET LUOT!", True, RED)
    msg2 = small_font.render("An phim bat ky ", True, WHITE)
    screen.blit(msg, (SCREEN_WIDTH//2 - msg.get_width()//2, 300))
    screen.blit(msg2, (SCREEN_WIDTH//2 - msg2.get_width()//2, 400))
    pygame.display.flip()

    waiting = True
    while waiting:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                hand_controller.stop_detection()
                pygame.quit(); sys.exit()
            elif event.type == pygame.KEYDOWN:
                waiting = False

